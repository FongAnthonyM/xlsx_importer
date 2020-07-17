#!/usr/bin/env python
# -*- coding: utf-8 -*-
""" xlsx_importer.py
Description: A module that reads excel files into an object structure and uses Pandas DataFrames to represent tabular
    data.
"""
__author__ = "Anthony Fong"
__copyright__ = "Copyright 2020, Anthony Fong"
__credits__ = ["Anthony Fong"]
__license__ = ""
__version__ = "1.0.0"
__maintainer__ = "Anthony Fong"
__email__ = ""
__status__ = "Prototype"

# Default Libraries #
import pathlib

# Downloaded Libraries #
import openpyxl
from openpyxl.utils import range_boundaries
import pandas
# import xlrd       need this one for pandas


# Local Libraries #


# Definitions #
# Classes #
class xlsxFile:
    """A class for reading excel files.

    Args:
        path (:obj:`str` or :obj:'Path', optional): Path of excel file.
        init (:obj:`bool`, optional): If the object automatically populates itself.

    Attributes:
        _path (:obj:`Path`): Path of excel file.
        op_workbook (:obj:`dict` of :obj:`Workbook`): An openpyxl object of the excel workbook.
        op_worksheets (:obj:`dict` of :obj:`Worksheet`): A dictionary of openpyxl worksheets with the names as keys
        op_tables (:obj:`dict` of :obj:`dict` of :obj:`str`): A dictionary of dictionaries with table
            ranges of each worksheets.
        worksheets (:obj:`dict` of :obj:`DataFrame`): A dictionary of Pandas dataframes containing
            the excel worksheets.
        tables (:obj:`dict` of :obj:`dict` of :obj:`DataFrame`): A dictionary of dictionaries with Pandas dataframes
            containing the table of each worksheet.
    """
    def __init__(self, path=None, init=True):
        if path is None:
            self._path = None
        else:
            self.path = path
        self.op_workbook = None
        self.op_worksheets = dict()
        self.op_tables = dict()

        self.worksheets = dict()
        self.tables = dict()

        if init:
            self.load()

    @property
    def path(self):
        """:obj:`Path`: Returns the _path attribute, a Path object of the path to the excel file.

        The setter takes a String or Path object and sets the _path attribute as a Path object or None if passed.
        """
        return self._path

    @path.setter
    def path(self, value):
        if isinstance(value, pathlib.Path) or value is None:
            self._path = value
        else:
            self._path = pathlib.Path(value)

    @property
    def sheet_names(self):
        """:obj:`list` of :obj: `string`: Returns a list of worksheet names from excel workbook."""
        return self.op_workbook.sheetnames

    def load(self, path=None):
        """Loads the data from the excel sheet into its corresponding attributes.

        Args:
            path (:obj:`str` or :obj:'Path', optional): Override current attribute path and the load from the new path.
        """
        self.load_wb(path=path)
        self.load_ws()
        self.load_tables()

    def load_wb(self, path=None):
        """Loads the data from the excel sheet as an openpyxl Workbook and assigns it to op_workbook.

        Args:
            path (:obj:`str` or :obj:'Path', optional): Override current attribute path and the load from the new path.
        """
        if path is not None:
            self.path = path
        self.op_workbook = openpyxl.load_workbook(self.path)

    def load_ws(self):
        """Loads the worksheets from op_workbook into op_worksheets and worksheets as a dictionaries where names
        of the worksheets are the keys.
        """
        for i, key in enumerate(self.sheet_names):
            self.op_worksheets[key] = self.op_workbook.worksheets[i]
            self.worksheets[key] = pandas.read_excel(self.path, index_col=None, header=None)

    def load_tables(self):
        """Loads the tables from all the worksheets into op_tables and table as dictionaries of dictionaries where
        the outer dictionary names worksheets and the inner dictionaries names tables.
        """
        op_tables = dict()
        tables = dict()
        for name in self.op_worksheets.keys():
            op_tables[name] = self.load_op_tables(name)
            tables[name] = self.load_pd_tables(name)
        self.op_tables = op_tables
        self.tables = tables

    def load_op_tables(self, name):
        """Loads the tables from the specified worksheet as a dictionary where the keys are the names of the tables and
        the values are ranges of the tables.

        Args:
            name (str): The name of the worksheet to get the tables from.

        Returns:
            :obj:`dict` of :obj:`str`: Contains the excel ranges of each of the tables in the worksheet.
        """
        tables = dict()
        for tbl_name, tbl_range in self.op_worksheets[name].tables.items():
            tables[tbl_name] = tbl_range
        return tables

    def load_pd_tables(self, name):
        """Loads the tables from the specified worksheet as a dictionary where the keys are the names of the tables and
        the values are Pandas Dataframes of the tables.

        Args:
            name (str): The name of the worksheet to get the tables from.

        Returns:
            :obj:`dict` of :obj:`str`: Contains the Pandas Dataframes of each of the tables in the worksheet.
        """
        tables = dict()
        for tbl_name, tbl_range in self.op_worksheets[name].tables.items():
            tables[tbl_name] = range2dataframe(self.op_worksheets[name], tbl_range)
        return tables


# Functions#
def range2dataframe(op_worksheet, range_):
    """Converts a range of an excel sheet into a Pandas dataframe.

    Args:
        op_worksheet (:obj:'Worksheet'): The worksheet which the range is located.
        range_ (str): The range to convert into a dataframe. It must be in the excel standard for a range.

    Returns:
        :obj:`DataFrame`: The Pandas dataframe of the excel range.
    """
    min_col, min_row, max_col, max_row = range_boundaries(range_)
    data = op_worksheet.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col, values_only=True)
    headers = next(data)
    return pandas.DataFrame(data, columns=headers)


# Main #
if __name__ == "__main__":
    db_path = pathlib.Path('/Users/changlab/Dropbox (UCSF Department of Neurological Surgery)/ChangLab/General Patient Info/EC223/')
    xfile = xlsxFile(db_path.joinpath('TrialNotes_EC223.xlsm'))
    other = xfile.tables["Tasks"]["PatientTasks"]

    print('done')

